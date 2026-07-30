"""
Excel Processor.

LIBRARY USED: openpyxl
-------------------------
openpyxl reads modern Excel files (.xlsx / .xlsm). It does NOT support the
legacy .xls binary format (that would require a different library, xlrd,
which has dropped .xls support in recent versions too) — this is a known,
documented limitation, not an oversight. If a client's company folder
contains an old .xls file, it will simply be skipped (not a supported
extension), which is safer than silently failing on a file we can't parse.

We open workbooks in `read_only=True` mode for lower memory use on large
spreadsheets, and `data_only=True` so formula cells return their last
CALCULATED value (e.g. 42) rather than the formula text itself
(e.g. "=SUM(A1:A10)") — for knowledge extraction we want the actual data,
not the formula.

WHAT GETS EXTRACTED
---------------------
Every sheet is extracted as its own labeled section. Within a sheet, each
non-empty row becomes one line, with cell values joined by " | ". Fully
empty rows are skipped to avoid bloating the text with blank lines.
"""

from pathlib import Path

import openpyxl

from modules.document_processors.base_processor import BaseDocumentProcessor, RawExtraction


class ExcelProcessor(BaseDocumentProcessor):
    """Extracts text and sheet count from .xlsx files."""

    document_type = "EXCEL"

    def _extract(self, file_path: Path) -> RawExtraction:
        try:
            workbook = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
        except Exception as exc:
            raise RuntimeError(f"Failed to open Excel file (it may be corrupted or password-protected): {exc}") from exc

        try:
            num_sheets = len(workbook.sheetnames)
            sheet_sections = []

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                row_lines = []
                for row in sheet.iter_rows(values_only=True):
                    cell_strings = [str(value).strip() for value in row if value is not None and str(value).strip() != ""]
                    if cell_strings:
                        row_lines.append(" | ".join(cell_strings))
                if row_lines:
                    sheet_sections.append(f"[Sheet: {sheet_name}]\n" + "\n".join(row_lines))
        finally:
            workbook.close()

        full_text = "\n\n".join(sheet_sections)
        return RawExtraction(text=full_text, num_pages=None, num_sheets=num_sheets)
